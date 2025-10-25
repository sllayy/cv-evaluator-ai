import subprocess
import streamlit as st
import json
import os
import shutil
import fitz  # PyMuPDF
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="CV Değerlendirme - Gereksinim Girişi")

st.title("🧠 CV Değerlendirme Sistemi")
st.subheader("İş Gereksinimlerini Giriş Ekranı")

# ----------------------------- FORM -----------------------------
pozisyon = st.text_input("Pozisyon")
beceriler = st.text_input("Aranan Beceriler (virgülle ayırın)")
min_deneyim = st.number_input("Minimum Deneyim (yıl)", 0, 20, 1)
egitim = st.selectbox("Eğitim Seviyesi", ["Lise", "Ön Lisans", "Lisans", "Yüksek Lisans", "Doktora"])
sertifikalar = st.text_input("Tercih Edilen Sertifikalar (virgülle ayırın)")
dil_ing = st.selectbox("İngilizce Seviyesi", ["A1", "A2", "B1", "B2", "C1", "C2"])
dil_2 = st.text_input("Diğer Dil ve Seviye (ör: İspanyolca:A1)")
linkedin = st.checkbox("LinkedIn profili isteniyor mu?", value=True)
projeler = st.checkbox("Proje deneyimi isteniyor mu?", value=True)

if st.button("Kaydet ve Başla"):
    data = {
        "pozisyon": pozisyon,
        "aranan_beceriler": [x.strip() for x in beceriler.split(",") if x.strip()],
        "minimum_deneyim_yili": min_deneyim,
        "egitim_seviyesi": egitim,
        "sertifikalar": [x.strip() for x in sertifikalar.split(",") if x.strip()],
        "dil": {"İngilizce": dil_ing},
        "ekstra": {"linkedin": linkedin, "proje_tecrubesi": projeler}
    }

    if ":" in dil_2:
        try:
            dil_adi, seviye = dil_2.split(":")
            data["dil"][dil_adi.strip()] = seviye.strip()
        except:
            st.warning("Diğer dil formatı geçersiz. Örn: İspanyolca:A1")

    os.makedirs("inputs", exist_ok=True)
    with open("inputs/job_requirements.json", "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

    st.success("✅ Gereksinimler başarıyla kaydedildi!")

# ----------------------------- CV YÜKLEME -----------------------------

st.subheader("📤 CV Dosyaları Yükle (.pdf/.docx)")

uploaded_files = st.file_uploader(
    "Birden fazla CV yükleyebilirsiniz", type=["pdf", "docx"], accept_multiple_files=True
)

def extract_name_from_pdf(file_path):
    try:
        with fitz.open(file_path) as doc:
            first_page = doc[0]
            blocks = first_page.get_text("blocks")
            blocks_sorted = sorted(blocks, key=lambda b: (b[1], b[0]))
            for block in blocks_sorted[:5]:
                text = block[4].strip()
                if 2 <= len(text.split()) <= 4 and text.replace(" ", "").isalpha() and text == text.upper():
                    return text.replace(" ", "_").title()
    except Exception as e:
        return None
    return None

if uploaded_files:
    os.makedirs("inputs/cvs", exist_ok=True)
    for uploaded_file in uploaded_files:
        extension = uploaded_file.name.split(".")[-1]
        temp_path = os.path.join("inputs/cvs", f"_temp_{uploaded_file.name}")
        with open(temp_path, "wb") as f:
            shutil.copyfileobj(uploaded_file, f)

        if extension.lower() == "pdf":
            name = extract_name_from_pdf(temp_path)
            new_filename = f"{name}.{extension}" if name else f"cv_{uploaded_file.name}"
        else:
            new_filename = f"cv_{uploaded_file.name}"

        final_path = os.path.join("inputs/cvs", new_filename)
        if os.path.exists(final_path):
            st.warning(f"⚠️ `{new_filename}` zaten yüklü, atlandı.")
            os.remove(temp_path)
            continue

        shutil.move(temp_path, final_path)
        st.success(f"✅ `{uploaded_file.name}` → `{os.path.basename(final_path)}` olarak kaydedildi.")

# ----------------------------- ANALİZ -----------------------------

st.subheader("📊 CV Analizi")

if st.button("Analizi Başlat"):
    with st.spinner("CV analizi yapılıyor..."):
        result = subprocess.run(["python", "main.py"], capture_output=True, text=True)

    if result.returncode == 0:
        st.success("✅ Analiz tamamlandı!")
        st.subheader("📝 Terminal Çıktısı")
        st.code(result.stdout)
    else:
        st.error("❌ Hata oluştu.")
        st.text(result.stderr)

# ----------------------------- RAPOR İNDİRME -----------------------------

st.subheader("📥 Excel Raporu")

if os.path.exists("outputs/scored_results.json"):
    with open("outputs/scored_results.json", "r", encoding="utf-8") as f:
        results = json.load(f)

    df = pd.DataFrame(results)
    
    #  "diller" sütununu kaldırdım
    if "diller" in df.columns:
        df = df.drop(columns=["diller"])

    excel_path = "outputs/cv_degerlendirme_raporu.xlsx"
    df.to_excel(excel_path, index=False)

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    if ws is not None:
        for col in ws.columns:
            max_length = 0
            column = col[0].column
            if column is not None:
                column_letter = get_column_letter(column)
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = max_length + 5
                ws.column_dimensions[column_letter].width = adjusted_width
        wb.save(excel_path)

    with open(excel_path, "rb") as f:
        st.download_button(
            label="📊 Excel Raporunu İndir",
            data=f,
            file_name="cv_degerlendirme_raporu.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

